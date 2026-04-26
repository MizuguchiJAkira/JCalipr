"""Excel export for a batch of measurement sets.

One row per fish, one column per measurement, plus metadata columns at the
front (fish_id, locality, date, image filename). We also emit a second sheet
with quality-control info: which landmarks were missing, calibration method
and confidence, and any warning notes from the ruler detector.

We use openpyxl directly (no pandas) to keep dependencies light.
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .measurement_engine import (
    MeasurementSet,
    measurement_column_order,
    measurement_labels,
)
from .ruler_calibration import CalibrationResult


DEFAULT_METADATA_COLUMNS: tuple[str, ...] = (
    "fish_id",
    "locality",
    "collection_date",
    "image_filename",
)


@dataclass
class ExportRecord:
    """One specimen's data packaged for export."""

    measurements: MeasurementSet
    # Calibrations per view — rendered on the QC sheet for provenance.
    calibrations: dict[str, CalibrationResult]
    image_filename: str = ""


def export_to_xlsx(
    records: Sequence[ExportRecord],
    output_path: str | Path,
    metadata_columns: Iterable[str] = DEFAULT_METADATA_COLUMNS,
) -> Path:
    """Write ``records`` to an xlsx workbook at ``output_path``.

    The workbook has two sheets:

    * ``Measurements`` — metadata columns + one column per measurement,
      with numeric values in mm / mm^2.
    * ``QC`` — calibration method / confidence / notes per view, plus a
      ``missing_landmarks`` column summarizing any gaps.

    Returns the resolved path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    meas_sheet = wb.active
    assert meas_sheet is not None
    meas_sheet.title = "Measurements"
    _write_measurements_sheet(meas_sheet, records, list(metadata_columns))

    qc_sheet = wb.create_sheet("QC")
    _write_qc_sheet(qc_sheet, records)

    wb.save(output_path)
    return output_path


def _write_measurements_sheet(
    sheet: Worksheet,
    records: Sequence[ExportRecord],
    metadata_columns: list[str],
) -> None:
    measurement_keys = measurement_column_order()
    labels = measurement_labels()

    header = [*metadata_columns, *(labels[k] for k in measurement_keys)]
    sheet.append(header)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    for col_idx in range(1, len(header) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        row: list[float | str] = []
        for col in metadata_columns:
            if col == "image_filename":
                row.append(rec.image_filename)
            elif col == "fish_id":
                row.append(rec.measurements.fish_id)
            else:
                row.append(rec.measurements.metadata.get(col, ""))
        for key in measurement_keys:
            v = rec.measurements.values.get(key)
            if v is None or math.isnan(v.value):
                row.append("")
            else:
                row.append(round(v.value, 3))
        sheet.append(row)

    # Reasonable column widths.
    for col_idx in range(1, len(header) + 1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = max(
            14, min(40, len(str(header[col_idx - 1])) + 2)
        )


def _write_qc_sheet(sheet: Worksheet, records: Sequence[ExportRecord]) -> None:
    header = [
        "fish_id",
        "image_filename",
        "view",
        "calibration_method",
        "px_per_mm",
        "confidence",
        "calibration_notes",
        "missing_landmarks",
    ]
    sheet.append(header)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    for col_idx in range(1, len(header) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for rec in records:
        missing = sorted(
            {
                lm
                for mv in rec.measurements.values.values()
                for lm in mv.missing_landmarks
            }
        )
        missing_str = ", ".join(missing) if missing else ""
        for view_name, calib in rec.calibrations.items():
            sheet.append(
                [
                    rec.measurements.fish_id,
                    rec.image_filename,
                    view_name,
                    calib.method,
                    round(calib.px_per_mm, 4),
                    round(calib.confidence, 3),
                    calib.notes,
                    missing_str,
                ]
            )

    for col_idx in range(1, len(header) + 1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = 18
